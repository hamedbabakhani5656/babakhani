import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def create_header_cell_style():
    header_style = {
        'font': Font(color="701830", bold=True),
        'fill': PatternFill(fill_type="solid", fgColor="D3D3D3"),
        'alignment': Alignment(horizontal="center", vertical="center"),
        'border': Border(left=Side(border_style="thin", color="000000"),
                         right=Side(border_style="thin", color="000000"),
                         top=Side(border_style="thin", color="000000"),
                         bottom=Side(border_style="thin", color="000000"))
    }
    return header_style

def create_cell_style():
    cell_style = {
        'font': Font(color="701830"),
        'alignment': Alignment(horizontal="center", vertical="center"),
        'border': Border(left=Side(border_style="dashed", color="701830"),
                         right=Side(border_style="dashed", color="701830"),
                         top=Side(border_style="dashed", color="701830"),
                         bottom=Side(border_style="dashed", color="701830"))
    }
    return cell_style

def create_delivered_cell_style():
    delivered_style = {
        'font': Font(color="701830"),
        'alignment': Alignment(horizontal="center", vertical="center"),
        'fill': PatternFill(fill_type="solid", fgColor="00FF00"),
        'border': Border(left=Side(border_style="thin", color="000000"),
                         right=Side(border_style="thin", color="000000"),
                         top=Side(border_style="thin", color="000000"),
                         bottom=Side(border_style="thin", color="000000"))
    }
    return delivered_style

def set_cell_border_style(merged_cell, sheet):
    thin_border = Border(left=Side(border_style="thin", color="000000"),
                         right=Side(border_style="thin", color="000000"),
                         top=Side(border_style="thin", color="000000"),
                         bottom=Side(border_style="thin", color="000000"))
    for row in sheet.iter_rows(min_row=merged_cell.min_row, max_row=merged_cell.max_row,
                               min_col=merged_cell.min_col, max_col=merged_cell.max_col):
        for cell in row:
            cell.border = thin_border

def set_border_style(merged_cell, sheet):
    thin_border = Border(left=Side(border_style="thin", color="000000"),
                         right=Side(border_style="thin", color="000000"),
                         top=Side(border_style="thin", color="000000"),
                         bottom=Side(border_style="thin", color="000000"))
    for row in sheet.iter_rows(min_row=merged_cell.min_row, max_row=merged_cell.max_row,
                               min_col=merged_cell.min_col, max_col=merged_cell.max_col):
        for cell in row:
            cell.border = thin_border

def get_old_query(start_date, end_date):
    query = """
    SELECT from_mobile_number, OPERATORS, smsstatus, SUM(farsia) AS farsi, SUM(latina) AS latin FROM (
        SELECT from_mobile_number, status AS smsstatus, SUM(farsis) AS farsia, SUM(latins) AS latina, prefix,
        (CASE
            WHEN prefix LIKE '934' THEN 'Kish'
            WHEN prefix LIKE '932' THEN 'Talia'
            ...
            ELSE 'MCI'
        END) AS OPERATORS
        FROM (
            SELECT from_mobile_number, status, dcs, SUBSTRING(dest_mobile_number from 3 for 3) AS prefix,
            SUBSTRING(dest_mobile_number from 3 for 5) AS prefix2, COUNT(*) AS farsis, 0 as latins
            FROM {table_name}
            WHERE dcs = 2 and status is not null
            AND date(creation_date) >= '{start_date}'
            AND date(creation_date) <= '{end_date}'
            GROUP BY from_mobile_number, status, dcs, 4, 5
            UNION ALL
            SELECT from_mobile_number, status, dcs, SUBSTRING(dest_mobile_number from 3 for 3) AS prefix,
            SUBSTRING(dest_mobile_number from 3 for 5) AS prefix2, 0 as farsis, COUNT(*) AS latins
            FROM {table_name}
            WHERE dcs <> 2 and status is not null
            AND date(creation_date) >= '{start_date}'
            AND date(creation_date) <= '{end_date}'
            GROUP BY from_mobile_number, status, dcs, 4, 5
            ORDER BY from_mobile_number, status
        )
        GROUP BY from_mobile_number, status, prefix, OPERATORS
    )
    GROUP BY from_mobile_number, OPERATORS, smsstatus
    ORDER BY from_mobile_number, smsstatus
    """.format(table_name="report", start_date=start_date, end_date=end_date)
    return query

def get_new_query(start_date, end_date):
    query = """
    SELECT from_mobile_number, OPERATORS, smsstatus, SUM(farsis) AS farsi, SUM(latins) AS latin FROM (
        SELECT from_mobile_number, status AS smsstatus, dcs, OPERATOR as OPERATORS, COUNT(*) AS farsis, 0 as latins
        FROM {table_name}
        WHERE dcs = 2 and status is not null
        AND date(creation_date) >= '{start_date}'
        AND date(creation_date) <= '{end_date}'
        GROUP BY from_mobile_number, smsstatus, dcs, OPERATOR
        UNION ALL
        SELECT from_mobile_number, status AS smsstatus, dcs, OPERATOR as OPERATORS, 0 as farsis, COUNT(*) AS latins
        FROM {table_name}
        WHERE dcs <> 2 and status is not null
        AND date(creation_date) >= '{start_date}'
        AND date(creation_date) <= '{end_date}'
        GROUP BY from_mobile_number, smsstatus, dcs, OPERATORS
        ORDER BY from_mobile_number, smsstatus
    )
    GROUP BY from_mobile_number, OPERATORS, smsstatus
    ORDER BY from_mobile_number, smsstatus
    """.format(table_name="report", start_date=start_date, end_date=end_date)
    return query