from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ایجاد workbook و worksheet
workbook = Workbook()
worksheet = workbook.active

# تعیین نام ستون‌ها
columns = ["نام عضو", "سرخط", "اپراتور", "نوع پیام", "Accepted", "Rejected", "Delivered", "Undelivered", "جمع کل"]

# افزودن نام ستون‌ها به worksheet
worksheet.append(columns)

# تنظیمات برای رنگ طوسی
fill = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")
font = Font(color="FFFFFF", bold=True)
alignment = Alignment(horizontal="center", vertical="center")

# اعمال تنظیمات بر روی سلول‌های سرستون
for cell in worksheet[1]:
    cell.fill = fill
    cell.font = font
    cell.alignment = alignment

# merge و درج متن
merge_range = "A1:I1"
worksheet.merge_cells(merge_range)
merged_cell = worksheet[merge_range.split(":")[0]]
merged_cell.value = "گزارش ارسال پیامک فروردین ماه شرکت خدمات انفورماتیک بر اساس شماره/اپراتور"
merged_cell.alignment = alignment


merge_range = "A2:A3"
worksheet.merge_cells(merge_range)
merged_cell = worksheet[merge_range.split(":")[0]]
merged_cell.value = "نام عضو"
merged_cell.alignment = alignment

merge_range = "B2:B3"
worksheet.merge_cells(merge_range)
merged_cell = worksheet[merge_range.split(":")[0]]
merged_cell.value = "سرخط"
merged_cell.alignment = alignment

merge_range = "C2:C3"
worksheet.merge_cells(merge_range)
merged_cell = worksheet[merge_range.split(":")[0]]
merged_cell.value = "اپراتور"
merged_cell.alignment = alignment

merge_range = "D2:D3"
worksheet.merge_cells(merge_range)
merged_cell = worksheet[merge_range.split(":")[0]]
merged_cell.value = "نوع پیام"
merged_cell.alignment = alignment

merge_range = "E2:I2"
worksheet.merge_cells(merge_range)
merged_cell = worksheet[merge_range.split(":")[0]]
merged_cell.value = "وضعیت ارسال"
merged_cell.alignment = alignment

merge_range = "E3:E3"
worksheet.merge_cells(merge_range)
merged_cell = worksheet[merge_range.split(":")[0]]
merged_cell.value = "Accepted"
merged_cell.alignment = alignment

merge_range = "F3:F3"
worksheet.merge_cells(merge_range)
merged_cell = worksheet[merge_range.split(":")[0]]
merged_cell.value = "Rejected"
merged_cell.alignment = alignment

merge_range = "G3:G3"
worksheet.merge_cells(merge_range)
merged_cell = worksheet[merge_range.split(":")[0]]
merged_cell.value = "Delivered"
merged_cell.alignment = alignment

merge_range = "H3:H3"
worksheet.merge_cells(merge_range)
merged_cell = worksheet[merge_range.split(":")[0]]
merged_cell.value = "Undelivered"
merged_cell.alignment = alignment

merge_range = "I3:I3"
worksheet.merge_cells(merge_range)
merged_cell = worksheet[merge_range.split(":")[0]]
merged_cell.value = "جمع کل"
merged_cell.alignment = alignment

# ذخیره فایل
file_name = "sms_report.xlsx"
workbook.save(filename=file_name)

