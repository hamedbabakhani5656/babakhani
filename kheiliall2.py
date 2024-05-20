import cx_Oracle
import pandas as pd
import openpyxl
from openpyxl import Workbook
from persiantools.jdatetime import JalaliDate
from datetime import datetime
from openpyxl.styles import PatternFill

# اطلاعات اتصال به دیتابیس
username = "hamed"
password = "hamed"
dsn = "orcl"  # نام دیتابیس مانند 'localhost/XEPDB1' در Oracle
host = "localhost"  # می‌تواند 'localhost' یا هاست دیگری باشد
port = "1521"  # پورت مانند '1521' در Oracle

# اتصال به دیتابیس
connection = cx_Oracle.connect(user=username, password=password, dsn=dsn, encoding="UTF-8")

# ایجاد یک کرسور برای اجرای کوئری‌ها
cursor = connection.cursor()

# تابع برای تبدیل تاریخ میلادی به شمسی
def miladi_to_shamsi(miladi_date):
    date_obj = datetime.strptime(miladi_date, '%Y-%m-%d')
    jalali_date = JalaliDate.to_jalali(date_obj.year, date_obj.month, date_obj.day)
    return jalali_date

# لیست نام ماه‌ها
month_names = [
    None, "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور",
    "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"
]



# درخواست تاریخ شروع از کاربر
start_date = input("لطفا تاریخ شروع (به فرمت YYYY-MM-DD) را وارد کنید: ")

# درخواست تاریخ پایان از کاربر
end_date = input("لطفا تاریخ پایان (به فرمت YYYY-MM-DD) را وارد کنید: ")

# تبدیل تاریخ‌های میلادی به شمسی
start_date_jalali = miladi_to_shamsi(start_date)
end_date_jalali = miladi_to_shamsi(end_date)




# کوئری اجرایی
query = """
SELECT
    'بانک صادرات' AS "نام عضو",
    from_mobile_number AS "سرخط",
    CASE
        WHEN operator = 'mci' THEN 'همراه اول'
        WHEN operator = 'samantel' THEN 'سامانتل'
        WHEN operator = 'irancell' THEN 'ایرانسل'
        WHEN operator = 'rightel' THEN 'رایتل'
        ELSE operator
    END AS "اپراتور",
    'فارسی' AS "نوع پیام",
    SUM(CASE WHEN status = 'SMSC_MESSAGE_ACCEPTED' THEN farsi ELSE 0 END) AS "Accepted",
    SUM(CASE WHEN status = 'SMSC_MESSAGE_REJECTED' THEN farsi ELSE 0 END) AS "Rejected",
    SUM(CASE WHEN status = 'SMSC_MESSAGE_DELIVERED' THEN farsi ELSE 0 END) AS "Delivered",
    SUM(CASE WHEN status = 'SMSC_MESSAGE_UNDELIVERABLE' THEN farsi ELSE 0 END) AS "Undelivered",
    SUM(farsi) AS "جمع کل"
FROM
    report
WHERE
    creation_date >= :start_date
    AND creation_date < :end_date
    AND operator IS NOT NULL
GROUP BY
    from_mobile_number,
    CASE
        WHEN operator = 'mci' THEN 'همراه اول'
        WHEN operator = 'samantel' THEN 'سامانتل'
        WHEN operator = 'irancell' THEN 'ایرانسل'
        WHEN operator = 'rightel' THEN 'رایتل'
        ELSE operator
    END

UNION

SELECT
    'بانک صادرات' AS "نام عضو",
    from_mobile_number AS "سرخط",
    CASE
        WHEN operator = 'mci' THEN 'همراه اول'
        WHEN operator = 'samantel' THEN 'سامانتل'
        WHEN operator = 'irancell' THEN 'ایرانسل'
        WHEN operator = 'rightel' THEN 'رایتل'
        ELSE operator
    END AS "اپراتور",
    'غیر فارسی' AS "نوع پیام",
    SUM(CASE WHEN status = 'SMSC_MESSAGE_ACCEPTED' THEN latin ELSE 0 END) AS "Accepted",
    SUM(CASE WHEN status = 'SMSC_MESSAGE_REJECTED' THEN latin ELSE 0 END) AS "Rejected",
    SUM(CASE WHEN status = 'SMSC_MESSAGE_DELIVERED' THEN latin ELSE 0 END) AS "Delivered",
    SUM(CASE WHEN status = 'SMSC_MESSAGE_UNDELIVERABLE' THEN latin ELSE 0 END) AS "Undelivered",
    SUM(latin) AS "جمع کل"
FROM
    report
WHERE
    creation_date >= :start_date
    AND creation_date < :end_date
    AND operator IS NOT NULL
GROUP BY
    from_mobile_number,
    CASE
        WHEN operator = 'mci' THEN 'همراه اول'
        WHEN operator = 'samantel' THEN 'سامانتل'
        WHEN operator = 'irancell' THEN 'ایرانسل'
        WHEN operator = 'rightel' THEN 'رایتل'
        ELSE operator
    END
"""

# پارامترهایی که برای کوئری مورد نیاز است
params = {
    'start_date': start_date,
    'end_date': end_date
}

# اجرای کوئری با استفاده از پارامترها
cursor.execute(query, params)

# گرفتن نتیجه کوئری
result = cursor.fetchall()

# تبدیل نتیجه به DataFrame
df = pd.DataFrame(result, columns=["نام عضو", "سرخط", "اپراتور", "نوع پیام", "Accepted", "Rejected", "Delivered", "Undelivered", "جمع کل"])

# ذخیره DataFrame در فایل Excel
# ایجاد نام فایل شمسی
file_name = f"bsi_sms_Magfa_SmsStatistics_{start_date_jalali.year}{start_date_jalali.month:02d}{start_date_jalali.day:02d}-{end_date_jalali.year}{end_date_jalali.month:02d}{end_date_jalali.day:02d}.xlsx"
df.to_excel(file_name, index=False)

# تغییر رنگ خط اول به طوسی
workbook = openpyxl.load_workbook(file_name)
worksheet = workbook.active
first_row = worksheet[1]
for cell in first_row:
    cell.fill = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")
workbook.save(file_name)

# بستن کرسور و اتصال
cursor.close()
connection.close()

print("نتیجه با موفقیت در فایل Excel ذخیره شد.")
