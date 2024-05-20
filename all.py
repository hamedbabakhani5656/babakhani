from datetime import datetime
from persiantools.jdatetime import JalaliDate
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# تابع برای تبدیل تاریخ میلادی به شمسی
def miladi_to_shamsi(miladi_date):
    date_obj = datetime.strptime(miladi_date, '%Y%m%d')
    jalali_date = JalaliDate.to_jalali(date_obj.year, date_obj.month, date_obj.day)
    return jalali_date

# لیست نام ماه‌ها
month_names = [
    None, "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور",
    "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"
]

# دریافت تاریخ میلادی از کاربر
start_date_miladi = input("Enter start date (YYYYMMDD): ")
end_date_miladi = input("Enter end date (YYYYMMDD): ")

# تبدیل تاریخ‌های میلادی به شمسی
start_date_jalali = miladi_to_shamsi(start_date_miladi)
end_date_jalali = miladi_to_shamsi(end_date_miladi)

# ایجاد نام فایل شمسی
file_name = f"bsi_sms_Magfa_SmsStatistics_{start_date_jalali.year}{start_date_jalali.month:02d}{start_date_jalali.day:02d}-{end_date_jalali.year}{end_date_jalali.month:02d}{end_date_jalali.day:02d}.xlsx"

# ایجاد workbook و worksheet
workbook = Workbook()
worksheet = workbook.active

# تعیین نام ستون‌ها
columns = ["نام عضو", "سرخط", "اپراتور", "نوع پیام", "Accepted", "Rejected", "Delivered", "Undelivered", "جمع کل"]

# افزودن نام ستون‌ها به worksheet
worksheet.append(columns)

# تنظیمات برای رنگ طوسی
fill = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")
font = Font(color="FFFFFF", bold=True)
alignment = Alignment(horizontal="center", vertical="center")

# تنظیمات border
border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

# اعمال تنظیمات بر روی سلول‌های سرستون
for cell in worksheet[1]:
    cell.fill = fill
    cell.font = font
    cell.alignment = alignment
    cell.border = border

# merge و درج متن
merge_ranges = [
    "A1:I1",
    "A2:A3", "B2:B3", "C2:C3", "D2:D3",
    "E2:I2",
    "E3:E3", "F3:F3", "G3:G3", "H3:H3", "I3:I3"
]

# نام ماه متناظر با ماه ورودی کاربر
start_month_name = month_names[start_date_jalali.month]
end_month_name = month_names[end_date_jalali.month]

# اعمال تغییرات بر اساس نام ماه
if start_month_name == end_month_name:
    merged_values = [
        f"گزارش ارسال پیامک {start_date_jalali.year} {start_month_name} ماه شرکت خدمات انفورماتیک بر اساس شماره/اپراتور",
        "نام عضو", "سرخط", "اپراتور", "نوع پیام", "وضعیت ارسال",
        "Accepted", "Rejected", "Delivered", "Undelivered", "جمع کل"
    ]
else:
    merged_values = [
        f"گزارش ارسال پیامک {start_date_jalali.year} {start_month_name} تا {end_date_jalali.year} {end_month_name} شرکت خدمات انفورماتیک بر اساس شماره/اپراتور",
        "نام عضو", "سرخط", "اپراتور", "نوع پیام", "وضعیت ارسال",
        "Accepted", "Rejected", "Delivered", "Undelivered", "جمع کل"
    ]

for merge_range, merged_value in zip(merge_ranges, merged_values):
    worksheet.merge_cells(merge_range)
    merged_cell = worksheet[merge_range.split(":")[0]]
    merged_cell.value = merged_value
    merged_cell.alignment = alignment
    merged_cell.border = border

# اعمال تنظیمات بر روی همه سلول‌ها
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
    for cell in row:
        cell.fill = fill
        cell.alignment = alignment
        cell.border = border

# ذخیره فایل
workbook.save(filename=file_name)

print(f"Excel file '{file_name}' has been created successfully.")
