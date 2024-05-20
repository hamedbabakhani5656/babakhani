from datetime import datetime
from persiantools.jdatetime import JalaliDate
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import cx_Oracle

# اطلاعات اتصال به دیتابیس
username = "hamed"
password = "hamed"
dsn = "orcl"  # نام دیتابیس مانند 'localhost/XEPDB1' در Oracle
host = "localhost"  # می‌تواند 'localhost' یا هاست دیگری باشد
port = "1521"  # پورت مانند '1521' در Oracle

# اتصال به دیتابیس
connection = cx_Oracle.connect(user=username, password=password, dsn=dsn, encoding="UTF-8")

# ایجاد یک کرسور برای اجرای کوئری‌ها
cursor = connection.cursor()

# تابع برای تبدیل تاریخ میلادی به شمسی
def miladi_to_shamsi(miladi_date):
    date_obj = datetime.strptime(miladi_date, '%Y-%m-%d')
    jalali_date = JalaliDate.to_jalali(date_obj.year, date_obj.month, date_obj.day)
    return jalali_date

# دریافت تاریخ میلادی از کاربر
start_date_miladi = input("Enter start date (YYYY-MM-DD): ")
end_date_miladi = input("Enter end date (YYYY-MM-DD): ")

# تبدیل تاریخ‌های میلادی به شمسی
start_date_jalali = miladi_to_shamsi(start_date_miladi)
end_date_jalali = miladi_to_shamsi(end_date_miladi)

# ایجاد نام فایل شمسی
file_name = f"bsi_sms_Magfa_SmsStatistics_{start_date_jalali.year}{start_date_jalali.month:02d}{start_date_jalali.day:02d}-{end_date_jalali.year}{end_date_jalali.month:02d}{end_date_jalali.day:02d}.xlsx"

# ایجاد workbook و worksheet
workbook = Workbook()
worksheet = workbook.active

# تعیین نام ستون‌ها
columns = ["نام عضو", "سرخط", "اپراتور", "نوع پیام", "Accepted", "Rejected", "Delivered", "Undelivered", "جمع کل"]

# افزودن نام ستون‌ها به worksheet
worksheet.append(columns)

# تنظیمات برای رنگ طوسی
fill = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")
font = Font(color="FFFFFF", bold=True)
alignment = Alignment(horizontal="center", vertical="center")

# تنظیمات border
border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

# اعمال تنظیمات بر روی سلول‌های سرستون
for index, cell in enumerate(worksheet[1]):
    cell.fill = fill
    cell.font = font
    cell.alignment = alignment
    cell.border = border
    # اعمال رنگ طوسی برای سه خط اول
    if index < 3:
        cell.fill = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")

# کوئری اجرایی
query = """
SELECT
    'بانک صادرات' AS "نام عضو",
    from_mobile_number AS "سرخط",
    CASE
        WHEN operator = 'mci' THEN 'همراه اول'
        WHEN operator = 'samantel' THEN 'سامانتل'
        WHEN operator = 'irancell' THEN 'ایرانسل'
        WHEN operator = 'rightel' THEN 'رایتل'
        ELSE operator
    END AS "اپراتور",
    'فارسی' AS "نوع پیام",
    SUM(CASE WHEN status = 'SMSC_MESSAGE_ACCEPTED' THEN farsi ELSE 0 END) AS "Accepted",
    SUM(CASE WHEN status = 'SMSC_MESSAGE_REJECTED' THEN farsi ELSE 0 END) AS "Rejected",
    SUM(CASE WHEN status = 'SMSC_MESSAGE_DELIVERED' THEN farsi ELSE 0 END) AS "Delivered",
    SUM(CASE WHEN status = 'SMSC_MESSAGE_UNDELIVERABLE' THEN farsi ELSE 0 END) AS "Undelivered",
    SUM(farsi) AS "جمع کل"
FROM
    report
WHERE
    creation_date >= :start_date
    AND creation_date < :end_date
    AND operator IS NOT NULL
GROUP BY
    from_mobile_number,
    CASE
        WHEN operator = 'mci' THEN 'همراه اول'
        WHEN operator = 'samantel' THEN 'سامانتل'
        WHEN operator = 'irancell' THEN 'ایرانسل'
        WHEN operator = 'rightel' THEN 'رایتل'
        ELSE operator
    END

UNION

SELECT
    'بانک صادرات' AS "نام عضو",
    from_mobile_number AS "سرخط",
    CASE
        WHEN operator = 'mci' THEN 'همراه اول'
        WHEN operator = 'samantel' THEN 'سامانتل'
        WHEN operator = 'irancell' THEN 'ایرانسل'
        WHEN operator = 'rightel' THEN 'رایتل'
        ELSE operator
    END AS "اپراتور",
    'غیر فارسی' AS "نوع پیام",
    SUM(CASE WHEN status = 'SMSC_MESSAGE_ACCEPTED' THEN latin ELSE 0 END) AS "Accepted",
    SUM(CASE WHEN status = 'SMSC_MESSAGE_REJECTED' THEN latin ELSE 0 END) AS "Rejected",
    SUM(CASE WHEN status = 'SMSC_MESSAGE_DELIVERED' THEN latin ELSE 0 END) AS "Delivered",
    SUM(CASE WHEN status = 'SMSC_MESSAGE_UNDELIVERABLE' THEN latin ELSE 0 END) AS "Undelivered",
    SUM(latin) AS "جمع کل"
FROM
    report
WHERE
    creation_date >= :start_date
    AND creation_date < :end_date
    AND operator IS NOT NULL
GROUP BY
    from_mobile_number,
    CASE
        WHEN operator = 'mci' THEN 'همراه اول'
        WHEN operator = 'samantel' THEN 'سامانتل'
        WHEN operator = 'irancell' THEN 'ایرانسل'
        WHEN operator = 'rightel' THEN 'رایتل'
        ELSE operator
    END
"""

# اجرای کوئری
data = cursor.execute(query, start_date=start_date_miladi, end_date=end_date_miladi).fetchall()

# merge و درج متن
merge_ranges = [
    "A1:I1",
    "A2:A3", "B2:B3", "C2:C3", "D2:D3",
    "E2:I2",
    "E3:E3", "F3:F3", "G3:G3", "H3:H3", "I3:I3"
]


# لیست نام ماه‌ها
month_names = [
    None, "فروردین", "اردیبهشت", "خرداد", "تیر", "مرداد", "شهریور",
    "مهر", "آبان", "آذر", "دی", "بهمن", "اسفند"
]


# نام ماه متناظر با ماه ورودی کاربر
start_month_name = month_names[start_date_jalali.month]
end_month_name = month_names[end_date_jalali.month]

# اعمال تغییرات بر اساس نام ماه
if start_month_name == end_month_name:
    merged_values = [
        f"گزارش ارسال پیامک {start_date_jalali.year} {start_month_name} ماه شرکت خدمات انفورماتیک بر اساس شماره/اپراتور",
        "نام عضو", "سرخط", "اپراتور", "نوع پیام", "وضعیت ارسال",
        "Accepted", "Rejected", "Delivered", "Undelivered", "جمع کل"
    ]
else:
    merged_values = [
        f"گزارش ارسال پیامک {start_date_jalali.year} {start_month_name} تا {end_date_jalali.year} {end_month_name} شرکت خدمات انفورماتیک بر اساس شماره/اپراتور",
        "نام عضو", "سرخط", "اپراتور", "نوع پیام", "وضعیت ارسال",
        "Accepted", "Rejected", "Delivered", "Undelivered", "جمع کل"
    ]

for merge_range, merged_value in zip(merge_ranges, merged_values):
    worksheet.merge_cells(merge_range)
    merged_cell = worksheet[merge_range.split(":")[0]]
    merged_cell.value = merged_value
    merged_cell.alignment = alignment
    merged_cell.border = border

# افزودن داده‌های گزارش به worksheet
for row_data in data:
    worksheet.append(row_data)

# اعمال تنظیمات بر روی همه سلول‌ها
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
    for cell in row:
        cell.fill = fill
        cell.alignment = alignment
        cell.border = border

# ذخیره فایل
workbook.save(filename=file_name)

print(f"Excel file '{file_name}' has been created successfully.")

# بستن کرسور و اتصال
cursor.close()
connection.close()