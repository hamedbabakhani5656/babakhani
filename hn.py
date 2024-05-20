from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ایجاد workbook و worksheet
workbook = Workbook()
worksheet = workbook.active

# تعیین نام ستون‌ها
columns = ["نام عضو", "سرخط", "اپراتور", "نوع پیام", "Accepted", "Rejected", "Delivered", "Undelivered", "جمع کل"]

# افزودن نام ستون‌ها به worksheet
worksheet.append(columns)

# تنظیمات برای رنگ طوسی
fill = PatternFill(start_color="7F7F7F", end_color="7F7F7F", fill_type="solid")
font = Font(color="FFFFFF", bold=True)
alignment = Alignment(horizontal="center", vertical="center")

# تنظیمات border
border = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

# اعمال تنظیمات بر روی سلول‌های سرستون
for cell in worksheet[1]:
    cell.fill = fill
    cell.font = font
    cell.alignment = alignment
    cell.border = border

# merge و درج متن
merge_ranges = [
    "A1:I1",
    "A2:A3", "B2:B3", "C2:C3", "D2:D3",
    "E2:I2",
    "E3:E3", "F3:F3", "G3:G3", "H3:H3", "I3:I3"
]

merged_values = [
    "گزارش ارسال پیامک فروردین ماه شرکت خدمات انفورماتیک بر اساس شماره/اپراتور",
    "نام عضو", "سرخط", "اپراتور", "نوع پیام", "وضعیت ارسال",
    "Accepted", "Rejected", "Delivered", "Undelivered", "جمع کل"
]

for merge_range, merged_value in zip(merge_ranges, merged_values):
    worksheet.merge_cells(merge_range)
    merged_cell = worksheet[merge_range.split(":")[0]]
    merged_cell.value = merged_value
    merged_cell.alignment = alignment
    merged_cell.border = border

# اعمال تنظیمات بر روی همه سلول‌ها
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
    for cell in row:
        cell.fill = fill
        cell.alignment = alignment
        cell.border = border

# ذخیره فایل
file_name = "sms_report.xlsx"
workbook.save(filename=file_name)
